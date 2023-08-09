import streamlit as st
import pandas as pd
import plotly.express as px
import io
import base64

buffer = io.BytesIO()

st.set_page_config(page_title='Site Balance')
st.title('Site Balance')
st.subheader("Instructions:")
st.subheader("In the first upload box, select the Canvas File. In the second upload box, select you downloaded file from Ops Tracker")
st.subheader("A File Name Site Automation Calculation will be generated! Click the download button to view the calculated data")

uploaded_file1 = st.file_uploader('Choose a XLSX File', type='xlsx')
uploaded_file2 = st.file_uploader('Choose another XLSX File', type='xlsx')


import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
#from openpyxl.chart import BarChart, Reference
from openpyxl.utils import FORMULAE
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import string
import re
from operator import *
import xlsxwriter
import xlwt
import os.path 


i = 1


excel_file = pd.read_excel(uploaded_file1, sheet_name='FinalWithSectors')
excel_file[['Site Name','Tech','Mgr']]
Manager = excel_file[['Mgr','Tech']]
Tech = excel_file[['Tech']]

excel_file1 = pd.read_excel(uploaded_file2)
excel_file1[['Site Tech Name','Site Mgr. Name','Weight Call Volume','Weight Drive Time','Weight Equipment','Weight Site Access','Weight Site Type']]
SiteTotal = excel_file1[['Site Tech Name','Site Mgr. Name','Weight Call Volume','Weight Drive Time','Weight Equipment','Weight Site Access','Weight Site Type']]

writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')

Manager_Count = excel_file['Mgr'].value_counts()
print(Manager_Count)
Total = excel_file.groupby(['Tech','Mgr'])['Mgr'].value_counts()
print(Total)
Tech_Count = excel_file['Tech'].value_counts()
print(Tech_Count)



df = pd.DataFrame(Manager_Count)
df2 = pd.DataFrame(Manager)
df3 = pd.DataFrame(Total)
df4 = pd.DataFrame(SiteTotal)

df.to_excel(writer,sheet_name="Manager")
#df2.to_excel(writer,sheet_name="Tech")
df3.to_excel(writer,sheet_name="Total")
df4.to_excel(writer,sheet_name="Site_Totals")

writer.close()
wb = load_workbook('test.xlsx')

sheet1 = wb["Manager"] 
#sheet2 = wb["Tech"]
sheet3 = wb["Total"]
sheet4 = wb["Site_Totals"]

sheet1.column_dimensions['A'].width = 30
sheet1.column_dimensions['B'].width = 30
sheet1.column_dimensions['C'].width = 30
sheet1.column_dimensions['D'].width = 40
sheet1.column_dimensions['E'].width = 40
sheet1.column_dimensions['F'].width = 40
sheet1.column_dimensions['G'].width = 40
sheet1.column_dimensions['H'].width = 45
sheet1.column_dimensions['J'].width = 40
sheet1.column_dimensions['K'].width = 40
sheet1.column_dimensions['L'].width = 40
sheet1.column_dimensions['M'].width = 40

sheet3.column_dimensions['N'].width = 40
sheet3.column_dimensions['B'].width = 30
sheet3.column_dimensions['C'].width = 30
sheet3.column_dimensions['D'].width = 30
sheet3.column_dimensions['E'].width = 30
sheet3.column_dimensions['F'].width = 40
sheet3.column_dimensions['G'].width = 40
sheet3.column_dimensions['H'].width = 55
sheet3.column_dimensions['J'].width = 40
sheet3.column_dimensions['K'].width = 40
sheet3.column_dimensions['L'].width = 40
sheet3.column_dimensions['M'].width = 40
sheet3.column_dimensions['I'].width = 40

sheet4.column_dimensions['A'].width = 20
sheet4.column_dimensions['B'].width = 20
sheet4.column_dimensions['C'].width = 20
sheet4.column_dimensions['D'].width = 20
sheet4.column_dimensions['E'].width = 20
sheet4.column_dimensions['F'].width = 20
sheet4.column_dimensions['G'].width = 20
sheet4.column_dimensions['H'].width = 20

sheet1['B12'] = ('=SUM(B2:B11)')
sheet1['A12'] = 'Total'
sheet1['A1'] = 'Manager'
sheet1['B1'] = 'Current Site Count'
sheet1['A17'] = 'Target Avg.'
sheet1['B16'] = 'New Sites Per Tech'
sheet1['C16'] = 'New Points per Tech'
sheet1['D1'] = 'Current Techs'
sheet1['C1'] = 'Current Total Points'
sheet1['E1'] = 'Current Sites per Tech'
sheet1['F1'] = 'Current Points per Tech'
sheet1['G1'] = 'Suggested Change in Sites per Mgr'
sheet1['H1'] = 'Suggested Change in Points per Mgr'
sheet1['J1'] = 'Additional Headcount Needed'
sheet1['K1'] = 'New Total Headcount'
sheet1['L1'] = 'New Balance of Sites per Mgr'
sheet1['M1'] = 'New Balance of Points per Mgr'
sheet4['H1'] = 'Site Ranking Score'
sheet1['D1'].font = Font(bold=True)
sheet1['C1'].font = Font(bold=True)
sheet1['E1'].font = Font(bold=True)
sheet1['F1'].font = Font(bold=True)
sheet1['G1'].font = Font(bold=True)
sheet1['H1'].font = Font(bold=True)
sheet1['A12'].font = Font(bold=True)
sheet1['B1'].font = Font(bold=True)
sheet1['A17'].font = Font(bold=True)
sheet1['B16'].font = Font(bold=True)
sheet1['C16'].font = Font(bold=True)
sheet1['J1'].font = Font(bold = True)
sheet1['K1'].font = Font(bold = True)
sheet1['L1'].font = Font(bold = True)
sheet1['M1'].font = Font(bold = True)
sheet1['J11'].font = Font(bold=True)


sheet3['D1'].font = Font(bold=True)
sheet3['E1'].font = Font(bold=True)
sheet3['F1'].font = Font(bold=True)
sheet3['G1'].font = Font(bold=True)
sheet3['H1'].font = Font(bold=True)
sheet3['J1'].font = Font(bold = True)
sheet3['K1'].font = Font(bold = True)
sheet3['L1'].font = Font(bold = True)
sheet3['I1'].font = Font(bold = True)

sheet1['A20'] = 'New Target Avg.'
sheet1['B19'] = 'New Sites Per Tech'
sheet1['C19'] = 'New Points per Tech'
sheet1['J11'] = 'Total Head Add Count'
sheet1['K11'] = '=SUM(J2:J10)'
sheet1['B19'].font = Font(bold = True)
sheet1['C19'].font = Font(bold = True)
sheet1['A20'].font = Font(bold=True)
sheet1['B23'].font = Font(bold=True)
sheet1['C23'].font = Font(bold=True)
sheet1['D23'].font = Font(bold=True)
sheet1['E23'].font = Font(bold=True)
sheet3.auto_filter.ref = 'A1:B7584'

#sheet1['C19'].font = Font(bold=True)

sheet3['B1'] = 'Manager'
sheet3['C1'] = 'Current Site Count'
#sheet2['E16'] = 'New Sites Per Tech'
#sheet2['F16'] = 'New Points per Tech'
sheet3['D1'] = 'Current Total Points'
sheet3['E1'] = 'Current Sites per Tech'
sheet3['F1'] = 'Current Points per Tech'
sheet3['G1'] = 'Suggested Change in Sites per Tech'
sheet3['H1'] = 'Suggested Change in Points per Tech'
sheet3['I1'] = 'Additional Sites Added'
sheet3['J1'] = 'New Total Site Count'
sheet3['K1'] = 'New Balance of Sites per Tech'
sheet3['L1'] = 'New Balance of Points per Tech'


sheet1['B17'] =('=B12/D12')
sheet1['D12'] = ('=SUM(D2:D11)')
sheet1['B20'] = ('=B12/SUM(K2:K10)')
sheet1['C8'] = '=SUM(Site_Totals!G2:G1289)'
sheet1['C12'] = '=SUM(C2:C11)'
sheet1['C17'] = '=C12/D12'
sheet1['C20'] = '=C12/SUM(K2:K10)'

sheet1['A24'] = '=_xlfn.UNIQUE(A2:A9)'




from_row = 2
to_row = 10

for i in range(from_row,to_row+1):
    sheet1[f"D{i}"] = f'=COUNTIF(Total!B2:B104,Manager!A{i})'
    sheet1[f"E{i}"] = f'=B{i}/D{i}'
    sheet1[f"G{i}"] = f'=(B17*D{i})-B{i}'
    sheet1[f"K{i}"] = f'=J{i}+D{i}'
    sheet1[f"L{i}"] = f'=(B20*K{i})-B{i}'
    sheet1[f"F{i}"] = f'=C{i}/D{i}'
    sheet1[f"H{i}"] = f'=(C17*D{i})-C{i}'
    sheet1[f"M{i}"] = f'=(C20*K{i})-C{i}'
    sheet1[f"C{i}"] = f'=SUMIF(Site_Totals!C2:C1289,Manager!A{i},Site_Totals!H2:H1289)'

region_row = 24
endregion_row = 31
sheet1['A24'] = '=A2'
sheet1['A25'] = '=A3'
sheet1['A26'] = '=A4'
sheet1['A27'] = '=A5'
sheet1['A28'] = '=A6'
sheet1['A29'] = '=A7'
sheet1['A30'] = '=A8'
sheet1['A31'] = '=A9'

sheet1['A23'] = 'Managers Target Avg.'
sheet1['B23'] = 'New Sites Per Tech'
sheet1['C23'] = 'New Points per Tech'
sheet1['D23'] = 'New Target Avg. Sites Per Tech'
sheet1['E23'] = 'New Target Avg. Points per Tech'

for i in range(region_row,endregion_row+1):
    sheet1[f"B{i}"]=f'=_xlfn.IFS(A{i}=A2,B2,A{i}=A3,B3,A{i}=A4,B4,A{i}=A5,B5,A{i}=A6,B6,A{i}=A7,B7,A{i}=A8,B8,A{i}=A9,B9)/_xlfn.IFS(A{i}=A2,D2,A{i}=A3,D3,A{i}=A4,D4,A{i}=A5,D5,A{i}=A6,D6,A{i}=A7,D7,A{i}=A8,D8,A{i}=A9,D9)'
    sheet1[f"C{i}"]=f'=_xlfn.IFS(A{i}=A2,C2,A{i}=A3,C3,A{i}=A4,C4,A{i}=A5,C5,A{i}=A6,C6,A{i}=A7,C7,A{i}=A8,C8,A{i}=A9,C9)/_xlfn.IFS(A{i}=A2,D2,A{i}=A3,D3,A{i}=A4,D4,A{i}=A5,D5,A{i}=A6,D6,A{i}=A7,D7,A{i}=A8,D8,A{i}=A9,D9)'
    sheet1[f"D{i}"]=f'=_xlfn.IFS(A{i}=A2,B2,A{i}=A3,B3,A{i}=A4,B4,A{i}=A5,B5,A{i}=A6,B6,A{i}=A7,B7,A{i}=A8,B8,A{i}=A9,B9)/_xlfn.IFS(A{i}=A2,K2,A{i}=A3,K3,A{i}=A4,K4,A{i}=A5,K5,A{i}=A6,K6,A{i}=A7,K7,A{i}=A8,K8,A{i}=A9,K9)'
    sheet1[f"E{i}"]=f'=_xlfn.IFS(A{i}=A2,C2,A{i}=A3,C3,A{i}=A4,C4,A{i}=A5,C5,A{i}=A6,C6,A{i}=A7,C7,A{i}=A8,C8,A{i}=A9,C9)/_xlfn.IFS(A{i}=A2,K2,A{i}=A3,K3,A{i}=A4,K4,A{i}=A5,K5,A{i}=A6,K6,A{i}=A7,K7,A{i}=A8,K8,A{i}=A9,K9)'

from_row2 = 2
to_row2 = 104

for i in range(from_row2,to_row2+1):
    sheet3[f"E{i}"] = f'=Manager!B12/Total!C{i}'
    sheet3[f"G{i}"] = f'=(Manager!B17*1)-Total!C{i}'
    sheet3[f"J{i}"] = f'=I{i}+C{i}'
    sheet3[F"K{i}"] = f'=(Manager!B20*1)-Total!J{i}'
    sheet3[f"D{i}"] = f'=SUMIF(Site_Totals!B2:B1289,Total!A{i},Site_Totals!H2:H1289)'
    sheet3[f"F{i}"] = f'=D{i}/_xlfn.IFS(B2=B{i},Manager!E8,Total!B4=B{i},Manager!E5,Total!B5=B{i},Manager!E3,B6=B{i},Manager!E6,Total!B7=B{i},Manager!E2,Total!B13=B{i},Manager!E7,Total!B17=B{i},Manager!E4,Total!B28=B{i},Manager!E9,Total!B22=B{i},Manager!E10)'
    sheet3[f"H{i}"] = f'=(Manager!C17*Total!D{i})-Total!C{i}'
    sheet3[f"L{i}"] = f'=(Manager!C20*K{i})-C{i}'

from_row3 = 2
to_row3 = 1289

for i in range(from_row3,to_row3+1):
    sheet4[F"H{i}"] = f'=AVERAGE(D{i}:G{i})'
sheet3.column_dimensions['A'].width = 20
sheet3.column_dimensions['B'].width = 20
sheet3.column_dimensions['C'].width = 20
sheet3.column_dimensions['D'].width = 20
sheet3['B106'] = 'Total'
sheet3['C106'] = '=SUM(C2:C105)'

#username = os.getlogin()
#save_path = f'C:\\Users\\{username}\\Desktop'
#name_of_file = 'Site Automation Calculation'
#completeName = os.path.join(save_path, name_of_file+".xlsx")

fill_cell = PatternFill(patternType='solid',fgColor='008081')
fill_cell2 = PatternFill(patternType='solid',fgColor='FF2400')
fill_cell3 = PatternFill(patternType='solid',fgColor='90EE90')
sheet1['A1'].fill = fill_cell
sheet1['B1'].fill = fill_cell
sheet1['C1'].fill = fill_cell
sheet1['D1'].fill = fill_cell
sheet1['E1'].fill = fill_cell
sheet1['F1'].fill = fill_cell
sheet1['G1'].fill = fill_cell
sheet1['H1'].fill = fill_cell
sheet1['J1'].fill = fill_cell3
sheet1['K1'].fill = fill_cell3
sheet1['L1'].fill = fill_cell3
sheet1['M1'].fill = fill_cell3
sheet1['A17'].fill = fill_cell2
sheet1['B16'].fill = fill_cell2
sheet1['C16'].fill = fill_cell2
sheet1['A20'].fill = fill_cell2
sheet1['B19'].fill = fill_cell2
sheet1['C19'].fill = fill_cell2
sheet1['A23'].fill = fill_cell
sheet1['B23'].fill = fill_cell
sheet1['C23'].fill = fill_cell
sheet1['D23'].fill = fill_cell2
sheet1['E23'].fill = fill_cell2

double = Side(border_style='thick',color='0A0A0A')
range = sheet1['A16':'D21']
range2 = sheet1['A23':'E32']
range3 = sheet1['K1':'M12']


for cell in range:
    for x in cell:
        x.border=Border(top=double,right=double,bottom=double)
for cell in range2:
    for x in cell:
        x.border=Border(top=double,right=double,bottom=double)
for cell in range3:
    for x in cell:
        x.border=Border(top=double,left=double,right=double,bottom=double)


wb.save('Site Automation Calculation.xlsx')
wb = Workbook('Site Automation Calculation.xlsx')

excel_file3 = pd.read_excel('Site Automation Calculation.xlsx',sheet_name='Manager')
excel_file4 = pd.read_excel('Site Automation Calculation.xlsx',sheet_name='Total')
excel_file5 = pd.read_excel('Site Automation Calculation.xlsx',sheet_name='Site_Totals')

with pd.ExcelWriter(buffer,engine='xlsxwriter') as writer:
    excel_file3.to_excel(writer,sheet_name='Manager',index=False)
    excel_file4.to_excel(writer,sheet_name='Total',index=False)
    excel_file5.to_excel(writer,sheet_name='Site_Totals',index=False)
    writer.close()
    st.download_button(label='📥 Download Current Result',data=buffer,file_name= 'Site Automation Calculation.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
